"""Excel report generator using openpyxl — financial reports, data exports, summaries."""

from __future__ import annotations

import io
import logging
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Brand colors
TESLE_GOLD = "D4A853"
TESLE_DARK = "1A1A2E"
TESLE_LIGHT = "F5F5F5"
TESLE_WHITE = "FFFFFF"
TESLE_BORDER = "E0E0E0"


class ExcelGenerator:
    """Generate Excel workbooks with formatting and charts."""

    def __init__(self) -> None:
        self._header_font = Font(name="Calibri", bold=True, color=TESLE_WHITE, size=11)
        self._header_fill = PatternFill(start_color=TESLE_DARK, end_color=TESLE_DARK, fill_type="solid")
        self._gold_fill = PatternFill(start_color=TESLE_GOLD, end_color=TESLE_GOLD, fill_type="solid")
        self._alt_fill = PatternFill(start_color=TESLE_LIGHT, end_color=TESLE_LIGHT, fill_type="solid")
        self._border = Border(
            left=Side(style="thin", color=TESLE_BORDER),
            right=Side(style="thin", color=TESLE_BORDER),
            top=Side(style="thin", color=TESLE_BORDER),
            bottom=Side(style="thin", color=TESLE_BORDER),
        )

    def generate(
        self,
        data: dict[str, Any],
        doc_type: str = "report",
        **options: Any,
    ) -> bytes:
        """Generate an Excel workbook."""
        wb = Workbook()

        if doc_type == "invoice":
            self._build_invoice(wb, data)
        elif doc_type == "report":
            self._build_report(wb, data)
        elif doc_type == "financial":
            self._build_financial(wb, data)
        else:
            self._build_generic(wb, data)

        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()
        buf.close()
        return raw

    # ── Invoice ──────────────────────────────────────────────────────────

    def _build_invoice(self, wb: Workbook, data: dict) -> None:
        ws = wb.active
        ws.title = "Invoice"

        company = data.get("company", {})
        ws.merge_cells("A1:F1")
        ws["A1"] = company.get("name", "TESLE")
        ws["A1"].font = Font(name="Calibri", bold=True, size=18, color=TESLE_GOLD)

        ws["A3"] = "Invoice #:"
        ws["B3"] = data.get("invoice_number", "INV-0000")
        ws["A4"] = "Date:"
        ws["B4"] = data.get("date", "N/A")
        ws["A5"] = "Due Date:"
        ws["B5"] = data.get("due_date", "N/A")
        ws["A6"] = "Status:"
        ws["B6"] = data.get("status", "pending").upper()

        for row in range(3, 7):
            ws.cell(row=row, column=1).font = Font(bold=True)

        # Bill To
        bill_to = data.get("bill_to", {})
        if bill_to:
            ws["D3"] = "Bill To:"
            ws["D3"].font = Font(bold=True)
            ws["D4"] = bill_to.get("name", "")
            if bill_to.get("address"):
                ws["D5"] = bill_to["address"]

        # Line items
        items = data.get("items", [])
        if items:
            headers = ["#", "Description", "Qty", "Unit Price", "Amount"]
            start_row = 8
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=i, value=h)
                cell.font = self._header_font
                cell.fill = self._header_fill
                cell.border = self._border
                cell.alignment = Alignment(horizontal="center")

            for idx, item in enumerate(items, 1):
                row = start_row + idx
                qty = item.get("quantity", 1)
                price = float(item.get("unit_price", item.get("price", 0)))
                ws.cell(row=row, column=1, value=idx).border = self._border
                ws.cell(row=row, column=2, value=item.get("description", item.get("name", ""))).border = self._border
                ws.cell(row=row, column=3, value=qty).border = self._border
                ws.cell(row=row, column=4, value=price).border = self._border
                ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=5, value=qty * price).border = self._border
                ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
                if idx % 2 == 0:
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).fill = self._alt_fill

            # Totals
            total_row = start_row + len(items) + 2
            subtotal = data.get("subtotal", 0)
            tax_rate = data.get("tax_rate", 0)
            total = data.get("total", 0)

            if not total and items:
                subtotal = sum(
                    float(it.get("quantity", 1)) * float(it.get("unit_price", it.get("price", 0)))
                    for it in items
                )
                total = subtotal + subtotal * (tax_rate / 100) if tax_rate else subtotal

            ws.cell(row=total_row, column=4, value="Subtotal:").font = Font(bold=True)
            ws.cell(row=total_row, column=5, value=subtotal).number_format = '"$"#,##0.00'

            if tax_rate:
                ws.cell(row=total_row + 1, column=4, value=f"Tax ({tax_rate}%):").font = Font(bold=True)
                ws.cell(row=total_row + 1, column=5, value=subtotal * tax_rate / 100).number_format = '"$"#,##0.00'

            ws.cell(row=total_row + 2, column=4, value="Total:").font = Font(bold=True, size=12)
            ws.cell(row=total_row + 2, column=5, value=total).number_format = '"$"#,##0.00'
            ws.cell(row=total_row + 2, column=5).font = Font(bold=True, size=12, color=TESLE_GOLD)

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

    # ── Report ───────────────────────────────────────────────────────────

    def _build_report(self, wb: Workbook, data: dict) -> None:
        # Summary sheet
        ws = wb.active
        ws.title = data.get("title", "Report")[:31]

        if data.get("title"):
            ws.merge_cells("A1:F1")
            ws["A1"] = data["title"]
            ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=TESLE_GOLD)

        row = 3
        if data.get("summary"):
            ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value=data["summary"])
            ws.merge_cells(f"A{row}:F{row}")
            row += 2

        # Data sheets
        tables = data.get("tables", [])
        for tbl in tables:
            sheet_name = tbl.get("title", "Data")[:31]
            sheet = wb.create_sheet(title=sheet_name)
            headers = tbl.get("headers", [])
            rows = tbl.get("rows", [])

            for i, h in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=i, value=h)
                cell.font = self._header_font
                cell.fill = self._header_fill
                cell.border = self._border
                cell.alignment = Alignment(horizontal="center")

            for r_idx, row_data in enumerate(rows, 2):
                for c_idx, val in enumerate(row_data, 1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = self._border
                    if r_idx % 2 == 0:
                        cell.fill = self._alt_fill

            # Auto-fit columns
            for col_idx in range(1, len(headers) + 1):
                max_len = max(
                    len(str(headers[col_idx - 1])),
                    max((len(str(row[col_idx - 1])) for row in rows), default=0),
                )
                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

            # Add chart if data is numeric
            if rows and len(headers) > 1:
                self._try_add_chart(sheet, headers, rows)

    def _try_add_chart(self, ws, headers: list, rows: list) -> None:
        """Add a chart if the data looks numeric."""
        try:
            numeric_cols = []
            for col_idx in range(2, len(headers)):
                try:
                    float(rows[0][col_idx - 1] if rows else 0)
                    numeric_cols.append(col_idx)
                except (ValueError, TypeError, IndexError):
                    pass

            if not numeric_cols:
                return

            chart = BarChart()
            chart.title = headers[numeric_cols[0]] if numeric_cols else ""
            chart.y_axis.title = "Value"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            cats = Reference(ws, min_col=1, min_row=2, max_row=min(len(rows) + 1, 50))
            for col_idx in numeric_cols[:3]:
                values = Reference(ws, min_col=col_idx + 1, min_row=1, max_row=min(len(rows) + 1, 50))
                chart.add_data(values, titles_from_data=True)

            chart.set_categories(cats)
            ws.add_chart(chart, f"A{len(rows) + 4}")
        except Exception:
            pass  # Charts are best-effort

    # ── Financial ────────────────────────────────────────────────────────

    def _build_financial(self, wb: Workbook, data: dict) -> None:
        """Build a multi-sheet financial report."""
        # Overview
        ws = wb.active
        ws.title = "Overview"
        if data.get("title"):
            ws.merge_cells("A1:E1")
            ws["A1"] = data["title"]
            ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=TESLE_GOLD)

        # KPIs
        kpis = data.get("kpis", [])
        if kpis:
            ws.cell(row=3, column=1, value="Key Metrics").font = Font(bold=True, size=12)
            headers = ["Metric", "Value", "Change", "Period"]
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=4, column=i, value=h)
                cell.font = self._header_font
                cell.fill = self._header_fill
                cell.border = self._border

            for idx, kpi in enumerate(kpis, 5):
                ws.cell(row=idx, column=1, value=kpi.get("name", "")).border = self._border
                ws.cell(row=idx, column=2, value=kpi.get("value", 0)).border = self._border
                ws.cell(row=idx, column=3, value=kpi.get("change", "")).border = self._border
                ws.cell(row=idx, column=4, value=kpi.get("period", "")).border = self._border

        # Data sheets
        tables = data.get("tables", [])
        for tbl in tables:
            sheet = wb.create_sheet(title=tbl.get("title", "Data")[:31])
            headers = tbl.get("headers", [])
            rows = tbl.get("rows", [])
            for i, h in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=i, value=h)
                cell.font = self._header_font
                cell.fill = self._header_fill
                cell.border = self._border
            for r_idx, row_data in enumerate(rows, 2):
                for c_idx, val in enumerate(row_data, 1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = self._border
                    if r_idx % 2 == 0:
                        cell.fill = self._alt_fill

    # ── Generic ──────────────────────────────────────────────────────────

    def _build_generic(self, wb: Workbook, data: dict) -> None:
        ws = wb.active
        ws.title = data.get("title", "Sheet 1")[:31]
        headers = data.get("headers", [])
        rows = data.get("rows", data.get("data", []))

        if headers:
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i, value=h)
                cell.font = self._header_font
                cell.fill = self._header_fill
                cell.border = self._border

        for r_idx, row_data in enumerate(rows, 2):
            if isinstance(row_data, dict):
                for c_idx, h in enumerate(headers, 1):
                    ws.cell(row=r_idx, column=c_idx, value=row_data.get(h, "")).border = self._border
            else:
                for c_idx, val in enumerate(row_data, 1):
                    ws.cell(row=r_idx, column=c_idx, value=val).border = self._border

        # Auto-fit columns
        for col_idx in range(1, max(len(headers), 1) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
